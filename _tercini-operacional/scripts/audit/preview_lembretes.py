"""Gera preview em Excel das 636 tarefas abertas com dueDateTime mas SEM lembrete.

Classifica cada uma em CRITICA (precisa de lembrete) ou ROTINEIRA (sem lembrete) e
sugere o tipo de lembrete (D-7, D-3, D-1) com horario.

Saida em audit_excel/06_preview_lembretes.xlsx para revisao do Paulo.
"""
import json
import re
from collections import Counter
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CACHE = json.loads(Path("/tmp/audit_cache.json").read_text())
TZ_BR = timezone(timedelta(hours=-3))
NOW = datetime.now(TZ_BR)
HOJE = NOW.date()

SAIDA = Path("/home/user/marketplaceterciniinss/audit_excel")
SAIDA.mkdir(exist_ok=True)

RE_ENTRY = re.compile(r"^(\d{2}\.\d{2}\.\d{4})\s*\(([A-Z])\)\s*:\s*(.*)$", re.MULTILINE)
RE_CPF11 = re.compile(r"#(\d{11})\b")


# Classificacao em ordem de prioridade (primeira regra que casa, vence)
REGRAS_CRITICAS = [
    # (tipo_critico, lembrete_dias_antes, hora, regex_no_body)
    ("Perícia médica", 1, "07:30", re.compile(r"per[íi]cia (m[ée]dica )?(agendada|marcada|judicial)|per[íi]cia em \d", re.IGNORECASE)),
    ("Audiência", 1, "07:30", re.compile(r"audi[êe]ncia (agendada|marcada|designada|em)|audi[êe]ncia.{0,30}\d{2}/\d{2}", re.IGNORECASE)),
    ("Avaliação social BPC", 1, "07:30", re.compile(r"avalia[çc][ãa]o social", re.IGNORECASE)),
    ("Manifestação sobre laudo", 3, "08:00", re.compile(r"manifestar.{0,30}(laudo|per[íi]cia)|manifesta[çc].{0,30}laudo", re.IGNORECASE)),
    ("Impugnação de laudo", 3, "08:00", re.compile(r"impugna[çc][ãa]o.{0,15}laudo|impugnar laudo", re.IGNORECASE)),
    ("Réplica processual", 3, "08:00", re.compile(r"r[ée]plica.{0,20}(protocolar|prazo|fazer)", re.IGNORECASE)),
    ("Embargos de declaração", 3, "08:00", re.compile(r"embargos.{0,20}declara", re.IGNORECASE)),
    ("Recurso CRPS prazo", 3, "08:00", re.compile(r"recurso (ordin[áa]rio|especial).{0,30}(crps|c[âa]mara|junta).{0,30}prazo|prazo.{0,15}recurso (ordin[áa]rio|crps)", re.IGNORECASE)),
    ("Mandado de Segurança - decadência", 7, "08:00", re.compile(r"\bMS\b.{0,30}120 dias|decad[êe]ncia.{0,15}MS|prazo.{0,15}120", re.IGNORECASE)),
    ("Apelação prazo", 3, "08:00", re.compile(r"apela[çc][ãa]o.{0,20}prazo|prazo.{0,15}apela", re.IGNORECASE)),
    ("Agravo prazo", 3, "08:00", re.compile(r"agravo.{0,20}prazo|prazo.{0,15}agravo", re.IGNORECASE)),
    ("Contestação prazo", 3, "08:00", re.compile(r"contesta[çc][ãa]o.{0,20}prazo|prazo.{0,15}contestar", re.IGNORECASE)),
    ("Cumprimento exigência INSS", 3, "08:00", re.compile(r"exig[êe]ncia.{0,20}(cumprir|prazo|cumprimento)|cumprir exig[êe]ncia", re.IGNORECASE)),
]

# Listas onde TODA tarefa com prazo merece lembrete
LISTAS_CRITICAS_AUTOMATICAS = ["🗓 Tarefas com Prazo"]

# Listas onde a maioria das tarefas com prazo NAO precisa lembrete (rotineiras)
LISTAS_ROTINEIRAS = ["🙏 Aposentadorias Futuras", "💵 Pagamentos"]


def parse_dt(d):
    if not d:
        return None
    try:
        s = d if isinstance(d, str) else d.get("dateTime", "")
        s = s.replace("Z", "+00:00")
        if "+" not in s and "-" not in s[10:]:
            s += "+00:00"
        return datetime.fromisoformat(s).astimezone(TZ_BR)
    except Exception:
        return None


def classificar(t, due_dt):
    """Retorna (categoria, tipo_critico, lembrete_dias_antes, hora_sugerida)."""
    body = t.get("body", "") or ""
    title = t.get("title", "") or ""
    lista = t["_lista"]
    contexto = f"{title}\n{body}"

    # Regra 1: lista que e' criticamente automatica
    if lista in LISTAS_CRITICAS_AUTOMATICAS:
        # Ainda assim, classificar pelo body se possivel
        for tipo, dias, hora, regex in REGRAS_CRITICAS:
            if regex.search(contexto):
                return ("CRITICA", tipo, dias, hora)
        return ("CRITICA", "Prazo processual generico", 3, "08:00")

    # Regra 2: matching por padrao no body
    for tipo, dias, hora, regex in REGRAS_CRITICAS:
        if regex.search(contexto):
            return ("CRITICA", tipo, dias, hora)

    # Regra 3: lista rotineira sem sinais criticos
    if lista in LISTAS_ROTINEIRAS:
        return ("ROTINEIRA", "Sem lembrete necessario", 0, "")

    # Regra 4: 🌻 INSS sem padrao critico = verificacao rotineira
    if lista == "🌻 INSS":
        return ("ROTINEIRA", "Verificacao rotineira no MEU INSS/PAT", 0, "")

    # Regra 5: 🙋 Escritório sem padrao critico = operacional
    if lista == "🙋 Escritório":
        return ("ROTINEIRA", "Tarefa operacional sem prazo processual", 0, "")

    # Default: marcar como INCERTA para revisao manual
    return ("INCERTA", "Sem padrao detectado - revisar manualmente", 0, "")


def estilizar_header(ws, n_cols):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def auto_largura(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            length = max(len(line) for line in str(val).split("\n"))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)


def main():
    todas = []
    for l in CACHE["listas"]:
        for t in l["tarefas"]:
            t["_lista"] = l["displayName"]
            todas.append(t)

    abertas = [t for t in todas if t.get("status") != "completed"]
    com_due_sem_reminder = []
    for t in abertas:
        if not t.get("dueDateTime"):
            continue
        if t.get("isReminderOn"):
            continue
        due_dt = parse_dt(t.get("dueDateTime"))
        if not due_dt:
            continue
        com_due_sem_reminder.append((due_dt, t))

    # Classificar
    classificadas = []
    for due_dt, t in com_due_sem_reminder:
        cat, tipo, dias, hora = classificar(t, due_dt)
        # Calcular data do lembrete
        if cat == "CRITICA" and dias > 0:
            data_lembrete = (due_dt - timedelta(days=dias)).date()
            datetime_lembrete = f"{data_lembrete.isoformat()} {hora}"
        else:
            datetime_lembrete = ""
        body = t.get("body", "") or ""
        m = RE_ENTRY.search(body)
        primeira_linha = body.split("\n")[0][:200] if body else ""
        ult_atualizacao = m.group(1) if m else "(sem entrada)"
        classificadas.append({
            "lista": t["_lista"],
            "title": t.get("title", ""),
            "due": due_dt.date().isoformat(),
            "categoria": cat,
            "tipo_critico": tipo,
            "lembrete_dias_antes": dias,
            "hora": hora,
            "datetime_lembrete": datetime_lembrete,
            "primeira_linha": primeira_linha,
            "ult_atualizacao": ult_atualizacao,
            "task_id": t["id"],
        })

    # Ordenar por categoria (CRITICA primeiro) e depois por data de due
    ordem_cat = {"CRITICA": 0, "INCERTA": 1, "ROTINEIRA": 2}
    classificadas.sort(key=lambda x: (ordem_cat.get(x["categoria"], 9), x["due"]))

    # Estatisticas
    cat_count = Counter(c["categoria"] for c in classificadas)
    tipo_count = Counter(c["tipo_critico"] for c in classificadas if c["categoria"] == "CRITICA")
    lista_critica_count = Counter(c["lista"] for c in classificadas if c["categoria"] == "CRITICA")
    lista_rotineira_count = Counter(c["lista"] for c in classificadas if c["categoria"] == "ROTINEIRA")

    wb = Workbook()

    # Aba 1 - Como funciona
    ws_intro = wb.active
    ws_intro.title = "Como funciona"
    ws_intro["A1"] = "PREVIEW DE LEMBRETES INTELIGENTES"
    ws_intro["A1"].font = Font(bold=True, size=14, color="1F4E78")
    intro = [
        "",
        f"Esta planilha lista as {len(classificadas)} tarefas ABERTAS que tem dueDateTime configurado mas NAO tem lembrete (isReminderOn=False).",
        "",
        "Cada uma foi classificada em uma de 3 categorias:",
        "",
        "  1. CRITICA - precisa de lembrete (perícia, audiência, prazo processual, exigência INSS, MS por mora)",
        "  2. INCERTA - padrão não detectado, revisar manualmente",
        "  3. ROTINEIRA - tarefas de verificação rotineira ou administrativa que NÃO precisam de lembrete",
        "",
        "Para CRITICA, o script propõe um LEMBRETE com:",
        "  - dias_antes da data de vencimento (D-1 para perícia/audiência, D-3 para manifestação/recurso, D-7 para MS)",
        "  - hora sugerida (07:30 para perícia, 08:00 para prazos processuais)",
        "",
        "INSTRUCOES:",
        "  1. Aba 'Lembretes propostos' - revisar cada linha",
        "  2. Coluna 'Aprovar?' - marcar SIM para aplicar, NAO para pular, EDITAR para mudar antes de aplicar",
        "  3. Coluna 'Observacao' - se EDITAR, indicar o que mudar (ex: 'D-2 em vez de D-3')",
        "",
        "Apos voce devolver a planilha, o script set_reminders_inteligente.py aplica via Microsoft Graph.",
        "",
        "Tarefas marcadas como ROTINEIRA por padrao NAO recebem lembrete (filosofia: reduzir ruido).",
        "Se quiser que alguma rotineira tenha lembrete tambem, marcar SIM na coluna 'Aprovar?' da linha correspondente.",
        "",
        "ESTATISTICAS:",
    ]
    for i, txt in enumerate(intro, 2):
        ws_intro.cell(row=i, column=1, value=txt)
    row = ws_intro.max_row + 1
    ws_intro.cell(row=row, column=1, value="").font = Font()
    row += 1
    ws_intro.cell(row=row, column=1, value=f"  Total {len(classificadas)} tarefas")
    row += 1
    for cat, n in cat_count.most_common():
        ws_intro.cell(row=row, column=1, value=f"  {cat}: {n}")
        row += 1
    row += 1
    ws_intro.cell(row=row, column=1, value="Tipos criticos detectados:")
    row += 1
    for tipo, n in tipo_count.most_common():
        ws_intro.cell(row=row, column=1, value=f"  {tipo}: {n}")
        row += 1
    row += 1
    ws_intro.cell(row=row, column=1, value="Distribuicao CRITICAS por lista:")
    row += 1
    for lista, n in lista_critica_count.most_common():
        ws_intro.cell(row=row, column=1, value=f"  {lista}: {n}")
        row += 1
    ws_intro.column_dimensions["A"].width = 110

    # Aba 2 - Lembretes propostos
    ws = wb.create_sheet("Lembretes propostos")
    headers = ["#", "Lista", "Título", "Due (BRT)", "Categoria", "Tipo crítico detectado",
               "Lembrete (D-X)", "Hora sugerida", "Datetime lembrete sugerido",
               "Última entrada", "Body (resumo)", "Aprovar?", "Observação"]
    ws.append(headers)
    estilizar_header(ws, len(headers))

    fill_critica = PatternFill("solid", fgColor="FFE699")
    fill_incerta = PatternFill("solid", fgColor="FFD9D9")
    fill_rotineira = PatternFill("solid", fgColor="E2EFDA")
    fill_aprovar = PatternFill("solid", fgColor="FFF2CC")

    for i, c in enumerate(classificadas, 1):
        if c["categoria"] == "CRITICA":
            row_fill = fill_critica
            aprovar_default = ""  # usuario decide
        elif c["categoria"] == "INCERTA":
            row_fill = fill_incerta
            aprovar_default = ""
        else:
            row_fill = fill_rotineira
            aprovar_default = "NAO"

        lembrete_label = f"D-{c['lembrete_dias_antes']}" if c['lembrete_dias_antes'] > 0 else "-"
        ws.append([
            i,
            c["lista"],
            c["title"],
            c["due"],
            c["categoria"],
            c["tipo_critico"],
            lembrete_label,
            c["hora"],
            c["datetime_lembrete"],
            c["ult_atualizacao"],
            c["primeira_linha"],
            aprovar_default,
            "",
        ])
        # Pintar linha pela categoria (apenas a coluna 5 categoria)
        ws.cell(row=i + 1, column=5).fill = row_fill
        ws.cell(row=i + 1, column=12).fill = fill_aprovar  # Aprovar?

    auto_largura(ws)

    # Aba 3 - Resumo por lista
    ws_lista = wb.create_sheet("Resumo por lista")
    ws_lista.append(["Lista", "CRITICA", "INCERTA", "ROTINEIRA", "Total"])
    estilizar_header(ws_lista, 5)
    listas_unicas = sorted(set(c["lista"] for c in classificadas))
    for lista in listas_unicas:
        cri = sum(1 for c in classificadas if c["lista"] == lista and c["categoria"] == "CRITICA")
        inc = sum(1 for c in classificadas if c["lista"] == lista and c["categoria"] == "INCERTA")
        rot = sum(1 for c in classificadas if c["lista"] == lista and c["categoria"] == "ROTINEIRA")
        total = cri + inc + rot
        ws_lista.append([lista, cri, inc, rot, total])
    # Totais
    total_cri = cat_count.get("CRITICA", 0)
    total_inc = cat_count.get("INCERTA", 0)
    total_rot = cat_count.get("ROTINEIRA", 0)
    ws_lista.append(["TOTAL", total_cri, total_inc, total_rot, total_cri + total_inc + total_rot])
    ws_lista.cell(row=ws_lista.max_row, column=1).font = Font(bold=True)
    auto_largura(ws_lista)

    out = SAIDA / "06_preview_lembretes.xlsx"
    wb.save(out)
    print(f"Excel salvo em {out}")
    print(f"  {len(classificadas)} tarefas analisadas")
    print(f"  CRITICAS: {total_cri}")
    print(f"  INCERTAS: {total_inc}")
    print(f"  ROTINEIRAS: {total_rot}")
    print(f"\nTipos criticos detectados:")
    for tipo, n in tipo_count.most_common():
        print(f"  {tipo}: {n}")


if __name__ == "__main__":
    main()
