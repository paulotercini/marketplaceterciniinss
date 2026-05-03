"""Gera 5 arquivos Excel para revisao manual do escritorio.

Saida em /home/user/marketplaceterciniinss/audit_excel/ (gitignored).

1. clientes_duplicados_mesma_lista.xlsx
2. top_carteira_pesada.xlsx
3. marcos_sem_cpf.xlsx
4. tarefas_fantasma_fora_marcos.xlsx
5. preview_auto_popular_dueDateTime.xlsx
"""
import json
import re
from collections import defaultdict, Counter
from datetime import datetime, date, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CACHE = json.loads(Path("/tmp/audit_cache.json").read_text())
TZ_BR = timezone(timedelta(hours=-3))
NOW = datetime.now(TZ_BR)
HOJE = NOW.date()
SAIDA = Path("/home/user/marketplaceterciniinss/audit_excel")
SAIDA.mkdir(exist_ok=True)

RE_ENTRY = re.compile(r"^(\d{2}\.\d{2}\.\d{4})\s*\(([A-Z])\)\s*:\s*(.*)$", re.MULTILINE)
RE_CPF11 = re.compile(r"#(\d{11})\b")
RE_NOME_CHAVE = re.compile(r"^(.*?)(?:\s+#\d+|\s*[★🤖🎂].*)?$")
RE_EM_DDMM = re.compile(r"em\s+(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)
RE_VERIFICAR_DDMM = re.compile(r"verificar.{0,30}em\s+(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)
RE_AGUARDAR_DDMM = re.compile(r"aguardar(?:\s+at[ée])?\s+(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)
RE_DDMM_VERIFICAR = re.compile(r"(\d{1,2})/(\d{1,2})(?:/(\d{4}))?\s*[,.]?\s*verificar", re.IGNORECASE)
RE_ATE_DDMM = re.compile(r"at[ée]\s+(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)
RE_PRORROG_DDMM = re.compile(r"prorrog.{0,30}(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)
RE_VENCE_DDMM = re.compile(r"vence.{0,15}(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)
RE_PRAZO_DDMM = re.compile(r"prazo.{0,15}(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)
RE_DCB_DDMM = re.compile(r"dcb.{0,15}(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)
RE_AUDIENCIA_DDMM = re.compile(r"audi[êe]ncia.{0,30}(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)
RE_PERICIA_DDMM = re.compile(r"per[íi]cia.{0,30}(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", re.IGNORECASE)


def estilizar_header(ws, n_cols):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def auto_largura(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            length = max(len(line) for line in str(val).split("\n"))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)


def chave_cliente(t):
    """Extrai chave do cliente: prefere CPF, fallback para nome normalizado."""
    title = t.get("title", "") or ""
    m = RE_CPF11.search(title)
    if m:
        return ("cpf", m.group(1))
    nome = re.sub(r"[★🤖🎂]+", "", title).strip()
    nome = re.sub(r"#\S+", "", nome).strip()
    nome_norm = nome.lower().strip()
    if nome_norm:
        return ("nome", nome_norm[:50])
    return None


def todas_tarefas():
    out = []
    for l in CACHE["listas"]:
        for t in l["tarefas"]:
            t["_lista"] = l["displayName"]
            out.append(t)
    return out


def parse_date_dt(s):
    try:
        return datetime.strptime(s, "%d.%m.%Y").date()
    except Exception:
        return None


# ==================== 1. CLIENTES DUPLICADOS NA MESMA LISTA ====================
def gerar_duplicados_mesma_lista():
    todas = todas_tarefas()
    # Considera TODAS as tarefas (abertas + concluidas) como Etapa 2 fez
    # Grupo apenas por CPF (compativel com a metrica 197 de Stage 2)

    grupos_cpf = defaultdict(list)
    for t in todas:
        chave = chave_cliente(t)
        if chave and chave[0] == "cpf":
            grupos_cpf[(t["_lista"], chave)].append(t)
    duplicados = {k: v for k, v in grupos_cpf.items() if len(v) > 1}

    # Tambem capturar duplicados por nome (sem CPF) - secundario
    grupos_nome = defaultdict(list)
    for t in todas:
        chave = chave_cliente(t)
        if chave and chave[0] == "nome":
            grupos_nome[(t["_lista"], chave)].append(t)
    duplicados_nome = {k: v for k, v in grupos_nome.items() if len(v) > 1}

    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicados mesma lista"

    headers = ["Lista", "Cliente (chave)", "Qtd no grupo", "Titulo", "Status",
               "Importance", "Due (BRT)", "Ult. atualizacao", "Ult. autor",
               "Tem CPF?", "Body length", "Suspeita?", "Observacao"]
    ws.append(headers)
    estilizar_header(ws, len(headers))

    casos_legitimos = 0
    casos_suspeitos = 0
    for (lista, chave), tarefas in sorted(duplicados.items()):
        chave_str = f"{chave[0]}={chave[1]}"
        for t in tarefas:
            body = t.get("body", "") or ""
            m = RE_ENTRY.search(body)
            ult_data = m.group(1) if m else "(sem entrada)"
            ult_autor = m.group(2) if m else "-"
            has_cpf = "Sim" if RE_CPF11.search(t.get("title", "")) else "Nao"
            due = t.get("dueDateTime", {})
            due_str = due.get("dateTime", "")[:10] if due else ""
            status = t.get("status", "")

            titles_no_grupo = [x.get("title", "") for x in tarefas]
            if len(set(titles_no_grupo)) == 1:
                suspeita = "SIM - titulos identicos"
                obs = "Provavel duplicata"
                casos_suspeitos += 1
            else:
                suspeita = "Nao - titulos diferentes"
                obs = "Provavel legitima"
                casos_legitimos += 1

            ws.append([
                lista,
                chave_str,
                len(tarefas),
                t.get("title", ""),
                status,
                t.get("importance", ""),
                due_str,
                ult_data,
                ult_autor,
                has_cpf,
                len(body),
                suspeita,
                obs,
            ])

    auto_largura(ws)

    # Aba: duplicados por NOME (sem CPF) - mostly Marcos
    ws_nome = wb.create_sheet("Duplicados por nome (Marcos)")
    headers_n = ["Lista", "Nome (chave)", "Qtd no grupo", "Titulo", "Status",
                 "Ult. atualizacao", "Ult. autor", "Body length", "Suspeita?"]
    ws_nome.append(headers_n)
    estilizar_header(ws_nome, len(headers_n))
    for (lista, chave), tarefas in sorted(duplicados_nome.items()):
        for t in tarefas:
            body = t.get("body", "") or ""
            m = RE_ENTRY.search(body)
            titles_no_grupo = [x.get("title", "") for x in tarefas]
            suspeita = "SIM" if len(set(titles_no_grupo)) == 1 else "Nao"
            ws_nome.append([
                lista, chave[1], len(tarefas), t.get("title", ""),
                t.get("status", ""),
                m.group(1) if m else "(sem entrada)",
                m.group(2) if m else "-",
                len(body), suspeita,
            ])
    auto_largura(ws_nome)

    # Aba resumo
    ws2 = wb.create_sheet("Resumo", 0)
    ws2.append(["Categoria", "Quantidade"])
    estilizar_header(ws2, 2)
    n_clientes_distintos_cpf = len(set(k[1] for k in duplicados.keys()))
    ws2.append(["Clientes (CPF) com duplicacao em alguma lista", n_clientes_distintos_cpf])
    ws2.append(["Grupos (cliente x lista) com duplicacao", len(duplicados)])
    ws2.append(["Tarefas envolvidas (CPF)", sum(len(v) for v in duplicados.values())])
    ws2.append(["Suspeita por titulo identico (CPF)", casos_suspeitos])
    ws2.append(["Legitimo por titulos diferentes (CPF)", casos_legitimos])
    ws2.append([])
    ws2.append(["Duplicados por NOME (sem CPF)", ""])
    ws2.append(["Grupos (nome x lista)", len(duplicados_nome)])
    ws2.append(["Tarefas envolvidas (nome)", sum(len(v) for v in duplicados_nome.values())])
    ws2.append([])
    ws2.append(["Por lista (CPF)", "Grupos"])
    por_lista = Counter(k[0] for k in duplicados.keys())
    for l, n in por_lista.most_common():
        ws2.append([l, n])
    auto_largura(ws2)

    out = SAIDA / "01_clientes_duplicados_mesma_lista.xlsx"
    wb.save(out)
    return out, len(duplicados), n_clientes_distintos_cpf, sum(len(v) for v in duplicados.values())


# ==================== 2. TOP CARTEIRA PESADA ====================
def gerar_carteira_pesada():
    todas = todas_tarefas()
    abertas = [t for t in todas if t.get("status") != "completed"]

    # Agrupar por chave_cliente, contar por lista
    cliente_dados = defaultdict(lambda: {"listas": Counter(), "titulos": [], "tarefas": []})
    for t in abertas:
        chave = chave_cliente(t)
        if chave and chave[0] == "cpf":  # Foco em clientes com CPF
            cliente_dados[chave]["listas"][t["_lista"]] += 1
            cliente_dados[chave]["titulos"].append(t.get("title", ""))
            cliente_dados[chave]["tarefas"].append(t)

    rankeados = sorted(cliente_dados.items(), key=lambda x: -sum(x[1]["listas"].values()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Carteira pesada"
    headers = ["Rank", "Cliente (CPF)", "Nome (titulo mais curto)", "Total tarefas",
               "Listas (qtd)", "Listas (detalhe)", "Diversidade beneficios"]
    ws.append(headers)
    estilizar_header(ws, len(headers))

    for rank, (chave, dados) in enumerate(rankeados[:30], 1):
        cpf = chave[1]
        nome_curto = min(dados["titulos"], key=len)
        # Limpar nome
        nome_limpo = re.sub(r"#\d+", "", nome_curto).strip()
        nome_limpo = re.sub(r"[★🤖🎂]", "", nome_limpo).strip()

        listas_detalhe = " | ".join(f"{l} ({n})" for l, n in dados["listas"].most_common())
        # Diversidade rapida - extrair palavras-chave
        body_concat = " ".join(t.get("body", "") for t in dados["tarefas"]).lower()
        beneficios = []
        for label, padrao in [
            ("idade", r"aposentadoria por idade"),
            ("TC", r"tempo de contribui|aposentadoria por tc"),
            ("especial", r"aposentadoria especial"),
            ("PCD", r"\bpcd\b|lc.?142"),
            ("B31", r"\bb31\b|aux[íi]lio[-\s]?doen"),
            ("B91", r"\bb91\b|invalidez|incapacidade permanente"),
            ("B94", r"\bb94\b|aux[íi]lio[-\s]?acidente"),
            ("BPC", r"\bbpc\b|loas|presta[çc][ãa]o continuada"),
            ("Pensao", r"pens[ãa]o por morte"),
        ]:
            if re.search(padrao, body_concat):
                beneficios.append(label)

        ws.append([
            rank,
            cpf,
            nome_limpo[:60],
            sum(dados["listas"].values()),
            len(dados["listas"]),
            listas_detalhe,
            ", ".join(beneficios),
        ])

    auto_largura(ws)
    out = SAIDA / "02_top_carteira_pesada.xlsx"
    wb.save(out)
    return out, len(rankeados)


# ==================== 3. MARCOS SEM CPF ====================
def gerar_marcos_sem_cpf():
    todas = todas_tarefas()
    abertas = [t for t in todas if t.get("status") != "completed"]

    marcos_sem = []
    for t in abertas:
        if t["_lista"] != "☕ Marcos":
            continue
        title = t.get("title", "") or ""
        if RE_CPF11.search(title):
            continue
        # Manter TODAS as tarefas sem CPF (incluindo administrativas) para Marcos decidir
        marcos_sem.append(t)

    wb = Workbook()
    ws = wb.active
    ws.title = "Marcos sem CPF"
    headers = ["#", "Titulo atual (so nome)", "CPF a adicionar", "Tem body?",
               "Body resumo (primeira entrada)", "Ultima atualizacao", "Ult. autor", "Observacao Marcos"]
    ws.append(headers)
    estilizar_header(ws, len(headers))

    for i, t in enumerate(sorted(marcos_sem, key=lambda x: x.get("title", "")), 1):
        body = t.get("body", "") or ""
        m = RE_ENTRY.search(body)
        primeira = body.split("\n")[0][:200] if body else ""
        ws.append([
            i,
            t.get("title", ""),
            "",  # CPF a preencher
            "Sim" if body.strip() else "Nao",
            primeira,
            m.group(1) if m else "-",
            m.group(2) if m else "-",
            "",  # Espaço para Marcos anotar
        ])

    # Destacar coluna a preencher
    fill_amarelo = PatternFill("solid", fgColor="FFF2CC")
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).fill = fill_amarelo  # CPF
        ws.cell(row=row, column=8).fill = fill_amarelo  # Observacao

    auto_largura(ws)

    # Aba instrucoes
    ws2 = wb.create_sheet("Instrucoes para Marcos", 0)
    ws2["A1"] = "INSTRUCOES"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
    instrucoes = [
        "",
        "Marcos, esta planilha lista todas as suas tarefas em '☕ Marcos' que NAO tem CPF no titulo.",
        "",
        "1. Para cada linha, preencha a coluna 'CPF a adicionar' com o CPF do cliente (somente digitos, 11 numeros).",
        "2. Se o item NAO for um cliente (e por exemplo um lembrete pessoal), deixe a coluna 'Observacao Marcos' marcada como 'NAO E CLIENTE'.",
        "3. Se voce nao se lembra do CPF, deixe a coluna em branco e marque 'BUSCAR' na observacao.",
        "",
        "Apos preencher, devolva a planilha. O sistema atualiza os titulos automaticamente para o padrao 'Nome #CPF11digitos'.",
        "",
        f"Total de tarefas a revisar: {len(marcos_sem)}",
    ]
    for i, txt in enumerate(instrucoes, 2):
        ws2.cell(row=i, column=1, value=txt)
    ws2.column_dimensions["A"].width = 100

    out = SAIDA / "03_marcos_sem_cpf.xlsx"
    wb.save(out)
    return out, len(marcos_sem)


# ==================== 4. TAREFAS FANTASMA FORA DE MARCOS ====================
def gerar_fantasmas_fora_marcos():
    todas = todas_tarefas()
    abertas = [t for t in todas if t.get("status") != "completed"]

    fantasmas = []
    for t in abertas:
        if t["_lista"] == "☕ Marcos":
            continue
        body = (t.get("body", "") or "").strip()
        cl = t.get("checklist", [])
        due = t.get("dueDateTime")
        if not body and not cl and not due:
            fantasmas.append(t)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fantasmas fora de Marcos"
    headers = ["#", "Lista", "Titulo", "Created", "Last modified", "Acao recomendada"]
    ws.append(headers)
    estilizar_header(ws, len(headers))

    for i, t in enumerate(sorted(fantasmas, key=lambda x: x["_lista"]), 1):
        ws.append([
            i,
            t["_lista"],
            t.get("title", ""),
            (t.get("createdDateTime", "") or "")[:10],
            (t.get("lastModifiedDateTime", "") or "")[:10],
            "",  # voce decide
        ])

    auto_largura(ws)

    ws2 = wb.create_sheet("Como agir", 0)
    ws2["A1"] = "TAREFAS FANTASMA FORA DA LISTA DO MARCOS"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
    sugestoes = [
        "",
        "Sao tarefas SEM body, SEM checklist e SEM prazo. Ficaram com so o titulo.",
        "",
        "Para cada uma, decidir uma das 4 acoes na coluna 'Acao recomendada':",
        "  - EXCLUIR. Tarefa perdeu o sentido, pode ir embora.",
        "  - PREENCHER. Vou abrir e adicionar contexto (CPF, telefone, body, prazo).",
        "  - MIGRAR para [lista]. Esta na lista errada.",
        "  - MANTER. E uma tarefa-lembrete de 1 linha que faz sentido (raro).",
        "",
        f"Total fora de Marcos: {len(fantasmas)}",
    ]
    for i, txt in enumerate(sugestoes, 2):
        ws2.cell(row=i, column=1, value=txt)
    ws2.column_dimensions["A"].width = 100

    out = SAIDA / "04_tarefas_fantasma_fora_marcos.xlsx"
    wb.save(out)
    return out, len(fantasmas)


# ==================== 5. PREVIEW AUTO-POPULAR dueDateTime ====================
def gerar_preview_due():
    todas = todas_tarefas()
    abertas = [t for t in todas if t.get("status") != "completed"]

    propostas = []
    PADROES = [
        (RE_PERICIA_DDMM, "pericia em DD/MM", 100),
        (RE_AUDIENCIA_DDMM, "audiencia em DD/MM", 100),
        (RE_DCB_DDMM, "DCB DD/MM", 95),
        (RE_VENCE_DDMM, "vence DD/MM", 90),
        (RE_PRAZO_DDMM, "prazo DD/MM", 90),
        (RE_VERIFICAR_DDMM, "verificar em DD/MM", 85),
        (RE_AGUARDAR_DDMM, "aguardar DD/MM", 85),
        (RE_PRORROG_DDMM, "prorrogacao DD/MM", 80),
        (RE_ATE_DDMM, "ate DD/MM", 75),
        (RE_DDMM_VERIFICAR, "DD/MM verificar", 70),
        (RE_EM_DDMM, "em DD/MM (generico)", 50),
    ]

    for t in abertas:
        if t.get("dueDateTime"):
            continue  # ja tem prazo
        body = t.get("body", "") or ""
        if not body:
            continue
        # Pegar a primeira entrada (mais recente). Se nao houver, pegar texto inteiro.
        m = RE_ENTRY.search(body)
        if m:
            texto_entrada = m.group(0)
            data_entrada = parse_date_dt(m.group(1))
            data_entrada_str = m.group(1)
            autor = m.group(2)
        else:
            texto_entrada = body[:500]
            data_entrada = None
            data_entrada_str = "(sem entrada datada)"
            autor = "-"

        # Buscar todos os padroes
        candidatas = []
        for regex, label, prioridade in PADROES:
            for mm in regex.finditer(texto_entrada):
                dia, mes, ano = mm.group(1), mm.group(2), mm.group(3) or ""
                try:
                    dia_i = int(dia)
                    mes_i = int(mes)
                    if dia_i < 1 or dia_i > 31 or mes_i < 1 or mes_i > 12:
                        continue
                    if ano:
                        candidata = date(int(ano), mes_i, dia_i)
                    else:
                        ano_ref = data_entrada.year if data_entrada else HOJE.year
                        candidata = date(ano_ref, mes_i, dia_i)
                        if data_entrada and candidata < data_entrada:
                            try:
                                candidata = date(ano_ref + 1, mes_i, dia_i)
                            except Exception:
                                continue
                except Exception:
                    continue

                # Filtros: aceitar somente datas razoaveis
                if candidata.year < 2024 or candidata.year > 2030:
                    continue
                if candidata < HOJE - timedelta(days=60):
                    continue  # passada ha mais de 60 dias - improvavel ser util
                candidatas.append((candidata, label, prioridade))

        if candidatas:
            # Preferir futuras com maior prioridade
            futuras = [c for c in candidatas if c[0] >= HOJE]
            if futuras:
                # Maior prioridade primeiro, depois data mais proxima
                futuras.sort(key=lambda x: (-x[2], x[0]))
                escolhida = futuras[0]
            else:
                candidatas.sort(key=lambda x: -x[2])
                escolhida = candidatas[0]

            propostas.append({
                "lista": t["_lista"],
                "title": t.get("title", ""),
                "data_entrada": data_entrada_str,
                "autor": autor,
                "trecho": texto_entrada[:300],
                "data_proposta": escolhida[0].isoformat(),
                "padrao": escolhida[1],
                "prioridade": escolhida[2],
                "task_id": t["id"],
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Preview auto-populate"
    headers = ["#", "Lista", "Titulo", "Ult. entrada (data)", "Autor", "Trecho usado",
               "Padrao detectado", "Confianca", "Data proposta para dueDateTime", "Aprovar?", "Observacao"]
    ws.append(headers)
    estilizar_header(ws, len(headers))

    for i, p in enumerate(sorted(propostas, key=lambda x: (x["data_proposta"], -x["prioridade"])), 1):
        ws.append([
            i,
            p["lista"],
            p["title"],
            p["data_entrada"],
            p["autor"],
            p["trecho"],
            p["padrao"],
            p["prioridade"],
            p["data_proposta"],
            "",  # SIM/NAO
            "",
        ])

    fill_verde = PatternFill("solid", fgColor="E2EFDA")
    fill_amarelo = PatternFill("solid", fgColor="FFF2CC")
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=9).fill = fill_verde
        ws.cell(row=row, column=10).fill = fill_amarelo
    auto_largura(ws)

    ws2 = wb.create_sheet("Como funciona", 0)
    ws2["A1"] = "PREVIEW DE AUTO-POPULAR dueDateTime"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
    txt = [
        "",
        "Esta planilha lista as tarefas ABERTAS que NAO TEM dueDateTime mas tem entrada datada no body com",
        "padrao de retorno futuro (em DD/MM, verificar em DD/MM, aguardar DD/MM).",
        "",
        "Para cada linha:",
        "  - Se a 'Data proposta' faz sentido, marque 'SIM' na coluna 'Aprovar?'",
        "  - Se NAO faz sentido, marque 'NAO' e na 'Observacao' explique (ex: data ja passou, e outro contexto).",
        "  - Se quiser uma data diferente, ponha em 'Observacao': nova data DD/MM/AAAA",
        "",
        "Apos voce devolver a planilha, o script aplica em massa via Microsoft Graph.",
        "",
        f"Total de propostas: {len(propostas)}",
        "",
        "FILTROS USADOS PARA INFERIR DATA:",
        "  - 'verificar em DD/MM'  - alta confianca",
        "  - 'aguardar DD/MM'  - alta confianca",
        "  - 'DD/MM, verificar'  - media confianca",
        "  - 'em DD/MM' generico  - menor confianca",
        "",
        "Caso o ano nao esteja explicito no body, eh inferido. Se a DD/MM ja passou em relacao a data da entrada,",
        "assume-se ano seguinte.",
    ]
    for i, t in enumerate(txt, 2):
        ws2.cell(row=i, column=1, value=t)
    ws2.column_dimensions["A"].width = 110

    out = SAIDA / "05_preview_auto_popular_dueDateTime.xlsx"
    wb.save(out)
    return out, len(propostas)


def main():
    print("Gerando arquivos Excel...\n")

    f1, n_grupos, n_clientes, n_tarefas = gerar_duplicados_mesma_lista()
    print(f"1. {f1.name}: {n_clientes} clientes, {n_grupos} grupos (cliente x lista), {n_tarefas} tarefas")

    f2, n_clientes = gerar_carteira_pesada()
    print(f"2. {f2.name}: {n_clientes} clientes ranqueados")

    f3, n_marcos = gerar_marcos_sem_cpf()
    print(f"3. {f3.name}: {n_marcos} tarefas para Marcos revisar")

    f4, n_fant = gerar_fantasmas_fora_marcos()
    print(f"4. {f4.name}: {n_fant} fantasmas fora de Marcos")

    f5, n_due = gerar_preview_due()
    print(f"5. {f5.name}: {n_due} propostas de dueDateTime")

    print(f"\nTodos em {SAIDA}")


if __name__ == "__main__":
    main()
